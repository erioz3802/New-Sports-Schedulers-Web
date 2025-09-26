# utils/bulk_processor.py - Bulk Upload Processing Engine
import openpyxl
from datetime import datetime, time, date
import re
from sqlalchemy import and_
from models.database import db
from models.league import League, Location
from models.game import Game, GameAssignment
from models.database import User
from utils.data_helpers import get_admin_leagues, get_all_locations, get_available_officials

# Replace these two functions in utils/bulk_processor.py:

def validate_upload_file(file_path):
    """Validate uploaded Excel file structure"""
    workbook = None
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook.active
        
        # Check if file has data
        if sheet.max_row < 2:
            return {'valid': False, 'error': 'File appears to be empty or has no data rows.'}
        
        # Get first row (headers)
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        # FIXED: Only truly required headers
        required_headers_games = [
            'League Name', 
            'Date (YYYY-MM-DD)', 
            'Time (HH:MM)', 
            'Location Name', 
            'Field/Court'
            # Game Level is NOT required - it's optional
        ]
        
        # Optional headers that are expected but not required
        optional_headers = [
            'Home Team',
            'Away Team', 
            'Game Level',        # This is optional!
            'Notes',
            'Special Instructions'
        ]
        
        # Check for required headers
        missing_headers = []
        for header in required_headers_games:
            if header not in headers:
                missing_headers.append(header)
        
        if missing_headers:
            return {
                'valid': False, 
                'error': f'Missing required columns: {", ".join(missing_headers)}. Found headers: {", ".join(headers)}'
            }
        
        # Determine template type
        template_type = 'with_assignments' if 'Official 1 Name' in headers else 'games_only'
        
        return {
            'valid': True, 
            'template_type': template_type,
            'total_rows': sheet.max_row - 1,  # Exclude header
            'headers_found': headers,
            'required_headers': required_headers_games,
            'optional_headers': optional_headers
        }
        
    except Exception as e:
        return {'valid': False, 'error': f'Error reading file: {str(e)}'}
    finally:
        if workbook:
            try:
                workbook.close()
            except:
                pass


def process_games_upload(file_path, admin_id, process_mode='save'):
    """
    Process uploaded games file
    
    Args:
        file_path: Path to uploaded Excel file
        admin_id: ID of admin uploading
        process_mode: 'validate_only', 'preview', or 'save'
    """
    
    # Initialize results
    results = {
        'success_count': 0,
        'error_count': 0,
        'warning_count': 0,
        'errors': [],
        'warnings': [],
        'preview_data': [],
        'process_mode': process_mode
    }
    
    workbook = None
    try:
        # ✅ FIXED: Use read_only=True and manual cleanup
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook.active
        
        # Create lookup dictionaries
        league_lookup = create_league_name_lookup(admin_id)
        location_lookup = create_location_name_lookup()
        official_lookup = None
        
        # Determine if this is an assignment template
        headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
        has_assignments = 'Official 1 Name' in headers
        
        if has_assignments:
            official_lookup = create_official_name_lookup(admin_id)
        
        # Get column mappings
        column_map = create_column_mapping(headers)
        
        # Process each data row
        for row_num in range(2, sheet.max_row + 1):
            try:
                # Extract row data
                row_data = extract_row_data(sheet, row_num, column_map)
                
                # Skip empty rows
                if is_empty_row(row_data):
                    continue
                
                # Validate and convert data
                validation_result = validate_and_convert_row(
                    row_data, row_num, league_lookup, location_lookup, official_lookup
                )
                
                if not validation_result['valid']:
                    results['errors'].extend(validation_result['errors'])
                    results['error_count'] += 1
                    continue
                
                game_data = validation_result['game_data']
                assignment_data = validation_result.get('assignment_data', [])
                
                # Check for conflicts
                conflict_warnings = check_game_conflicts(game_data, admin_id)
                if conflict_warnings:
                    results['warnings'].extend([f"Row {row_num}: {w}" for w in conflict_warnings])
                    results['warning_count'] += len(conflict_warnings)
                
                # Preview mode - just collect data
                if process_mode in ['validate_only', 'preview']:
                    preview_item = create_preview_item(game_data, assignment_data, row_num)
                    results['preview_data'].append(preview_item)
                
                # Save mode - create actual game
                elif process_mode == 'save':
                    game = create_game_from_data(game_data, admin_id)
                    
                    # Create assignments if present
                    if assignment_data and has_assignments:
                        create_assignments_from_data(game.id, assignment_data)
                    
                    results['success_count'] += 1
                else:
                    results['success_count'] += 1
                
            except Exception as e:
                error_msg = f"Row {row_num}: Unexpected error - {str(e)}"
                results['errors'].append(error_msg)
                results['error_count'] += 1
        
        # Commit database changes if in save mode
        if process_mode == 'save' and results['success_count'] > 0:
            db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        results['errors'].append(f"File processing error: {str(e)}")
        results['error_count'] += 1
    
    finally:
        # ✅ FIXED: Always close workbook
        if workbook:
            try:
                workbook.close()
            except:
                pass
    
    return results
def create_league_name_lookup(admin_id):
    """Create dictionary mapping league names to IDs for admin"""
    leagues = get_admin_leagues(admin_id)
    return {league['name']: league['id'] for league in leagues}

def create_location_name_lookup():
    """Create dictionary mapping location names to IDs"""
    locations = get_all_locations()
    return {location['name']: location['id'] for location in locations}

def create_official_name_lookup(admin_id):
    """Create dictionary mapping official names to IDs"""
    officials = get_available_officials(admin_id)
    lookup = {}
    for official in officials:
        full_name = f"{official['first_name']} {official['last_name']}"
        lookup[full_name] = official['id']
    return lookup

def create_column_mapping(headers):
    """Create mapping of column names to indices"""
    column_map = {}
    for idx, header in enumerate(headers):
        if header:
            column_map[header] = idx
    return column_map

def extract_row_data(sheet, row_num, column_map):
    """Extract data from a row based on column mapping"""
    row_data = {}
    row = sheet[row_num]
    
    for column_name, col_idx in column_map.items():
        if col_idx < len(row):
            cell_value = row[col_idx].value
            if cell_value is not None:
                row_data[column_name] = str(cell_value).strip()
            else:
                row_data[column_name] = ''
        else:
            row_data[column_name] = ''
    
    return row_data

def is_empty_row(row_data):
    """Check if row is empty (all required fields are empty)"""
    required_fields = ['League Name', 'Date (YYYY-MM-DD)', 'Location Name']
    return all(not row_data.get(field, '').strip() for field in required_fields)

def validate_and_convert_row(row_data, row_num, league_lookup, location_lookup, official_lookup):
    """Validate and convert row data to database format"""
    
    errors = []
    game_data = {}
    assignment_data = []
    
    # Validate League Name
    league_name = row_data.get('League Name', '').strip()
    if not league_name:
        errors.append(f"Row {row_num}: League Name is required")
    elif league_name not in league_lookup:
        errors.append(f"Row {row_num}: League '{league_name}' not found or not accessible")
    else:
        game_data['league_id'] = league_lookup[league_name]
    
    # Validate Date - ENHANCED VERSION
    date_str = row_data.get('Date (YYYY-MM-DD)', '').strip()
    if not date_str:
        errors.append(f"Row {row_num}: Date is required")
    else:
        try:
            game_date = None
            
            # Handle Excel date objects (when Excel stores as datetime)
            if hasattr(date_str, 'date'):
                game_date = date_str.date()
            elif isinstance(date_str, date):
                game_date = date_str
            elif isinstance(date_str, datetime):
                game_date = date_str.date()
            else:
                # Handle string dates - try multiple formats
                date_formats = [
                    '%Y-%m-%d %H:%M:%S',  # 2025-09-25 00:00:00 (FIRST PRIORITY)
                    '%Y-%m-%d',           # 2025-12-31
                    '%m/%d/%Y',           # 12/31/2025
                    '%m-%d-%Y',           # 12-31-2025
                    '%d/%m/%Y',           # 31/12/2025 (European)
                    '%Y/%m/%d',           # 2025/12/31
                    '%m/%d/%y',           # 12/31/25
                    '%d-%m-%Y',           # 31-12-2025
                    '%Y.%m.%d',           # 2025.12.31
                    '%m.%d.%Y',           # 12.31.2025
                ]
                
                # Remove any extra whitespace and convert to string
                date_str_clean = str(date_str).strip()
                
                # Try parsing with each format
                for date_format in date_formats:
                    try:
                        parsed_datetime = datetime.strptime(date_str_clean, date_format)
                        game_date = parsed_datetime.date()
                        break
                    except ValueError:
                        continue
                
                # If no format worked, try to handle Excel serial dates
                if game_date is None:
                    try:
                        # Excel stores dates as numbers (days since 1900-01-01)
                        if date_str_clean.replace('.', '').isdigit():
                            excel_date = float(date_str_clean)
                            # Convert Excel serial date to Python date
                            from datetime import timedelta
                            excel_epoch = date(1899, 12, 30)  # Adjusted for Excel bug
                            game_date = excel_epoch + timedelta(days=int(excel_date))
                    except (ValueError, OverflowError):
                        pass
            
            if game_date is None:
                errors.append(f"Row {row_num}: Invalid date format '{date_str}'. Supported formats: YYYY-MM-DD, MM/DD/YYYY, MM-DD-YYYY, etc.")
            else:
                game_data['date'] = game_date
                    
        except Exception as e:
            errors.append(f"Row {row_num}: Error parsing date '{date_str}': {str(e)}")
    
    # Validate Time - ENHANCED VERSION
    time_str = row_data.get('Time (HH:MM)', '').strip()
    if not time_str:
        errors.append(f"Row {row_num}: Time is required")
    else:
        try:
            game_time = None
            
            # Handle Excel time objects
            if isinstance(time_str, time):
                game_time = time_str
            elif hasattr(time_str, 'time'):
                game_time = time_str.time()
            elif isinstance(time_str, datetime):
                game_time = time_str.time()
            else:
                # Handle string times - try multiple formats
                time_str_clean = str(time_str).strip()
                
                # Remove common time suffixes
                time_str_clean = time_str_clean.upper().replace('AM', '').replace('PM', '').strip()
                
                # Try different time formats
                if ':' in time_str_clean:
                    time_parts = time_str_clean.split(':')
                    if len(time_parts) >= 2:
                        try:
                            hour = int(time_parts[0])
                            minute = int(time_parts[1])
                            
                            # Handle 12-hour format conversion if needed
                            if 'PM' in str(time_str).upper() and hour != 12:
                                hour += 12
                            elif 'AM' in str(time_str).upper() and hour == 12:
                                hour = 0
                            
                            if 0 <= hour <= 23 and 0 <= minute <= 59:
                                game_time = time(hour, minute)
                        except ValueError:
                            pass
                
                # Try Excel decimal time format (0.5 = 12:00)
                if game_time is None:
                    try:
                        if '.' in time_str_clean or time_str_clean.replace('.', '').isdigit():
                            decimal_time = float(time_str_clean)
                            if 0 <= decimal_time <= 1:
                                # Convert decimal to hours and minutes
                                total_minutes = int(decimal_time * 24 * 60)
                                hour = total_minutes // 60
                                minute = total_minutes % 60
                                game_time = time(hour, minute)
                    except ValueError:
                        pass
                
                # Try HHMM format (1430 = 14:30)
                if game_time is None and len(time_str_clean) == 4 and time_str_clean.isdigit():
                    try:
                        hour = int(time_str_clean[:2])
                        minute = int(time_str_clean[2:])
                        if 0 <= hour <= 23 and 0 <= minute <= 59:
                            game_time = time(hour, minute)
                    except ValueError:
                        pass
            
            if game_time is None:
                errors.append(f"Row {row_num}: Invalid time format '{time_str}'. Use HH:MM format (e.g., 14:30 or 2:30 PM)")
            else:
                game_data['time'] = game_time
                
        except Exception as e:
            errors.append(f"Row {row_num}: Error parsing time '{time_str}': {str(e)}")
    
    # Validate Location Name
    location_name = row_data.get('Location Name', '').strip()
    if not location_name:
        errors.append(f"Row {row_num}: Location Name is required")
    elif location_name not in location_lookup:
        errors.append(f"Row {row_num}: Location '{location_name}' not found. Add it in Location Management first.")
    else:
        game_data['location_id'] = location_lookup[location_name]
    
    # Optional text fields (teams are now optional)
    game_data['field_name'] = row_data.get('Field/Court', '').strip()
    game_data['home_team'] = row_data.get('Home Team', '').strip()
    game_data['away_team'] = row_data.get('Away Team', '').strip()
    
    # Optional fields
    game_data['game_level'] = row_data.get('Game Level', '').strip()
    game_data['notes'] = row_data.get('Notes', '').strip()
    game_data['special_instructions'] = row_data.get('Special Instructions', '').strip()
    
    # Process official assignments if present
    if official_lookup:
        for i in range(1, 4):  # Official 1, 2, 3
            official_name = row_data.get(f'Official {i} Name', '').strip()
            official_position = row_data.get(f'Official {i} Position', '').strip()
            
            if official_name:
                if official_name not in official_lookup:
                    errors.append(f"Row {row_num}: Official '{official_name}' not found or not accessible")
                else:
                    assignment_data.append({
                        'official_id': official_lookup[official_name],
                        'position': official_position or 'Official'
                    })
    
    return {
        'valid': len(errors) == 0,
        'errors': errors,
        'game_data': game_data,
        'assignment_data': assignment_data
    }
def check_game_conflicts(game_data, admin_id):
    """Check for potential scheduling conflicts"""
    warnings = []
    
    if 'date' not in game_data or 'time' not in game_data or 'location_id' not in game_data:
        return warnings
    
    # Check for same location/date/time conflicts
    existing_game = Game.query.filter(
        and_(
            Game.date == game_data['date'],
            Game.time == game_data['time'],
            Game.location_id == game_data['location_id'],
            Game.field_name == game_data.get('field_name', '')
        )
    ).first()
    
    if existing_game:
        warnings.append(f"Potential conflict: Same location/field/time already scheduled (Game ID: {existing_game.id})")
    
    return warnings

def create_preview_item(game_data, assignment_data, row_num):
    """Create preview item for display"""
    # Get names for display
    league = League.query.get(game_data.get('league_id'))
    location = Location.query.get(game_data.get('location_id'))
    
    preview = {
        'row': row_num,
        'league_name': league.name if league else 'Unknown',
        'date': game_data.get('date'),
        'time': game_data.get('time'),
        'location_name': location.name if location else 'Unknown',
        'field_name': game_data.get('field_name', ''),
        'home_team': game_data.get('home_team', ''),
        'away_team': game_data.get('away_team', ''),
        'assignments': []
    }
    
    # Add assignment previews
    for assignment in assignment_data:
        official = User.query.get(assignment['official_id'])
        if official:
            preview['assignments'].append({
                'official_name': f"{official.first_name} {official.last_name}",
                'position': assignment['position']
            })
    
    return preview

def create_game_from_data(game_data, admin_id):
    """Create Game object from validated data"""
    
    # Get league to determine fee
    league = League.query.get(game_data['league_id'])
    
    # Handle optional team names - get from game_data, not undefined variables
    home_team = game_data.get('home_team', '') or 'TBD'
    away_team = game_data.get('away_team', '') or 'TBD'
    
    game = Game(
        league_id=game_data['league_id'],
        location_id=game_data['location_id'],
        date=game_data['date'],
        time=game_data['time'],
        field_name=game_data.get('field_name', ''),
        home_team=home_team,
        away_team=away_team,
        level=game_data.get('game_level', ''),
        status='draft',
        fee_per_official=league.game_fee if league else 0.0,
        notes=game_data.get('notes', ''),
        special_instructions=game_data.get('special_instructions', ''),
        estimated_duration=120
    )
    
    db.session.add(game)
    db.session.flush()  # Get the ID
    
    return game

def create_assignments_from_data(game_id, assignment_data):
    """Create GameAssignment objects from assignment data"""
    
    for assignment in assignment_data:
        game_assignment = GameAssignment(
            game_id=game_id,
            official_id=assignment['official_id'],
            position=assignment['position'],
            status='assigned',
            assigned_at=datetime.utcnow()
        )
        db.session.add(game_assignment)