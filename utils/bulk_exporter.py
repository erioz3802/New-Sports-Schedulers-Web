# utils/bulk_exporter.py - Export Existing Games to Excel Format
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
import os
from datetime import datetime, date
from utils.data_helpers import get_admin_games
from models.game import Game, GameAssignment
from models.database import db

def export_admin_games(admin_id, league_id=None, date_from=None, date_to=None, include_assignments=False):
    """Export admin's games to Excel format matching import templates"""
    
    # Parse date filters
    date_from_obj = None
    date_to_obj = None
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Get games data
    games_data = get_admin_games(
        admin_id=admin_id,
        league_id=league_id,
        date_from=date_from_obj,
        date_to=date_to_obj
    )
    
    # Create workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    if include_assignments:
        sheet.title = "Games with Assignments Export"
        headers = [
            'League Name', 'Date (YYYY-MM-DD)', 'Time (HH:MM)', 'Location Name',
            'Field/Court', 'Home Team', 'Away Team', 'Game Level',
            'Official 1 Name', 'Official 1 Position',
            'Official 2 Name', 'Official 2 Position', 
            'Official 3 Name', 'Official 3 Position',
            'Notes', 'Special Instructions', 'Status'
        ]
    else:
        sheet.title = "Games Export"
        headers = [
            'League Name', 'Date (YYYY-MM-DD)', 'Time (HH:MM)', 'Location Name',
            'Field/Court', 'Home Team', 'Away Team', 'Game Level',
            'Notes', 'Special Instructions', 'Status'
        ]
    
    # Style headers
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        # Set column widths
        column_letter = openpyxl.utils.get_column_letter(col)
        if 'Name' in header and 'Official' not in header:
            sheet.column_dimensions[column_letter].width = 20
        elif 'Official' in header and 'Name' in header:
            sheet.column_dimensions[column_letter].width = 18
        elif header in ['Home Team', 'Away Team']:
            sheet.column_dimensions[column_letter].width = 16
        elif header in ['Notes', 'Special Instructions']:
            sheet.column_dimensions[column_letter].width = 25
        else:
            sheet.column_dimensions[column_letter].width = 15
    
    # Write data
    for row_num, game in enumerate(games_data, 2):
        col = 1
        
        # Basic game data
        sheet.cell(row=row_num, column=col).value = game['league_name']
        col += 1
        
        sheet.cell(row=row_num, column=col).value = game['date'].strftime('%Y-%m-%d') if game['date'] else ''
        col += 1
        
        sheet.cell(row=row_num, column=col).value = game['time'].strftime('%H:%M') if game['time'] else ''
        col += 1
        
        sheet.cell(row=row_num, column=col).value = game['location_name']
        col += 1
        
        sheet.cell(row=row_num, column=col).value = game['field_name'] or ''
        col += 1
        
        sheet.cell(row=row_num, column=col).value = game['home_team']
        col += 1
        
        sheet.cell(row=row_num, column=col).value = game['away_team']
        col += 1
        
        sheet.cell(row=row_num, column=col).value = game['game_level'] or ''
        col += 1
        
        # Assignment data if requested
        if include_assignments:
            assignments = game.get('assignments', [])
            
            # Fill up to 3 official slots
            for i in range(3):
                if i < len(assignments):
                    assignment = assignments[i]
                    sheet.cell(row=row_num, column=col).value = assignment['official_name']
                    col += 1
                    sheet.cell(row=row_num, column=col).value = assignment['position']
                    col += 1
                else:
                    # Empty official slots
                    sheet.cell(row=row_num, column=col).value = ''
                    col += 1
                    sheet.cell(row=row_num, column=col).value = ''
                    col += 1
        
        # Notes and status
        sheet.cell(row=row_num, column=col).value = game['notes'] or ''
        col += 1
        
        sheet.cell(row=row_num, column=col).value = game['special_instructions'] or ''
        col += 1
        
        sheet.cell(row=row_num, column=col).value = game['status'].title()
    
    # Add summary sheet
    create_export_summary_sheet(workbook, games_data, admin_id, include_assignments)
    
    # Save to temporary file
    temp_dir = tempfile.mkdtemp()
    
    # Create filename
    export_type = "with_assignments" if include_assignments else "games_only"
    date_range = ""
    if date_from_obj and date_to_obj:
        date_range = f"_{date_from_obj.strftime('%Y%m%d')}_to_{date_to_obj.strftime('%Y%m%d')}"
    elif date_from_obj:
        date_range = f"_from_{date_from_obj.strftime('%Y%m%d')}"
    elif date_to_obj:
        date_range = f"_to_{date_to_obj.strftime('%Y%m%d')}"
    
    filename = f"Games_Export_{export_type}{date_range}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = os.path.join(temp_dir, filename)
    workbook.save(file_path)
    
    return file_path, filename

def create_export_summary_sheet(workbook, games_data, admin_id, include_assignments):
    """Create summary sheet with export information"""
    
    summary_sheet = workbook.create_sheet("Export Summary", 0)
    
    # Title
    summary_sheet['A1'] = "Sports Scheduler - Games Export Summary"
    summary_sheet['A1'].font = Font(size=16, bold=True, color="366092")
    
    # Export details
    row = 3
    summary_sheet[f'A{row}'] = "Export Details:"
    summary_sheet[f'A{row}'].font = Font(bold=True)
    row += 1
    
    summary_sheet[f'A{row}'] = f"• Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    row += 1
    
    summary_sheet[f'A{row}'] = f"• Total Games: {len(games_data)}"
    row += 1
    
    summary_sheet[f'A{row}'] = f"• Export Type: {'With Assignments' if include_assignments else 'Games Only'}"
    row += 1
    
    summary_sheet[f'A{row}'] = f"• Admin ID: {admin_id}"
    row += 2
    
    # Games breakdown by status
    status_counts = {}
    league_counts = {}
    
    for game in games_data:
        status = game['status']
        league = game['league_name']
        
        status_counts[status] = status_counts.get(status, 0) + 1
        league_counts[league] = league_counts.get(league, 0) + 1
    
    if status_counts:
        summary_sheet[f'A{row}'] = "Games by Status:"
        summary_sheet[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for status, count in status_counts.items():
            summary_sheet[f'A{row}'] = f"• {status.title()}: {count}"
            row += 1
        row += 1
    
    if league_counts:
        summary_sheet[f'A{row}'] = "Games by League:"
        summary_sheet[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for league, count in league_counts.items():
            summary_sheet[f'A{row}'] = f"• {league}: {count}"
            row += 1
        row += 1
    
    # Instructions
    summary_sheet[f'A{row}'] = "Usage Instructions:"
    summary_sheet[f'A{row}'].font = Font(bold=True)
    row += 1
    
    instructions = [
        "• This file can be modified and re-imported using the bulk upload feature",
        "• Maintain the exact column structure when making changes",
        "• Use dropdown values when re-importing (download fresh template for current options)",
        "• The 'Status' column is for reference only and will be ignored during import",
        "• Remove this summary sheet before uploading if making changes"
    ]
    
    for instruction in instructions:
        summary_sheet[f'A{row}'] = instruction
        row += 1
    
    # Set column width
    summary_sheet.column_dimensions['A'].width = 80

def export_template_for_league(admin_id, league_id, include_assignments=False):
    """Export a template pre-filled with a specific league's information"""
    
    from utils.bulk_template_generator import generate_games_only_template, generate_games_with_assignments_template
    from models.league import League
    
    # Get league info
    league = League.query.get(league_id)
    if not league:
        raise ValueError(f"League {league_id} not found")
    
    # Generate base template
    if include_assignments:
        file_path, filename = generate_games_with_assignments_template(admin_id)
    else:
        file_path, filename = generate_games_only_template(admin_id)
    
    # Modify template to pre-select league
    workbook = openpyxl.load_workbook(file_path)
    main_sheet = workbook.active
    
    # Pre-fill league name in sample row and additional rows
    for row in range(2, 12):  # Fill first 10 rows with league name
        main_sheet.cell(row=row, column=1).value = league.name
    
    # Update filename to include league name
    league_safe_name = "".join(c for c in league.name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    new_filename = filename.replace('Template_', f'Template_{league_safe_name}_')
    
    # Save updated template
    new_file_path = file_path.replace(filename, new_filename)
    workbook.save(new_file_path)
    
    # Clean up original file
    os.unlink(file_path)
    
    return new_file_path, new_filename